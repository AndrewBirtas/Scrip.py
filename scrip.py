#!/usr/bin/python3
import sys
from openpyxl import load_workbook


def getFOF(FOF):
	wb = load_workbook(filename=f'/home/arch/Desktop/Master/{sys.argv[1]}')
	sheet = wb['Flows']

	line =[]

	for i in range(9, sheet.max_row + 1):
		line.append(sheet[i][10].value)
		line.append(sheet[i][18].value)
		line.append(sheet[i][19].value)
		if None in line:
			continue
		else:
			FOF.append(line)
		line = []



def concatenate(FOF, n):
	for element in range(0,n-1):
		print("\n")
		print(FOF[element])
		print("==========================================")
		try:
			for element2 in range(element+1, n):
				print(FOF[element2])
				if FOF[element][0] == FOF[element2][0]:
					if FOF[element][1] == FOF[element2][1]:
						print("CONCATENATE")
						FOF[element][2] = FOF[element][2] + " " + FOF[element2][2]
						print(FOF[element][2])
						del FOF[element2]
						print(len(FOF))
						concatenate(FOF, len(FOF))
						break
					else:
						if FOF[element][2] == FOF[element2][2]:
							print("CONCATENATE")
							FOF[element][1] = FOF[element][1] + " " + FOF[element2][1]
							print(FOF[element][1])
							del FOF[element2]
							print(len(FOF))
							concatenate(FOF, len(FOF))
							break
				else:
					if FOF[element][1] == FOF[element2][1] and FOF[element][2] == FOF[element2][2]:
						print("CONCATENATE")
						FOF[element][0] = FOF[element][0] + " " + FOF[element2][0]
						print(FOF[element][2])
						del FOF[element2]
						print(len(FOF))
						concatenate(FOF, len(FOF))
						break
		except IndexError:
			break



def showFOFDone(FOF):
	print("\n\n\n\n")
	for element in range(0, len(FOF)):
		print("\n\nSource:")
		x = FOF[element][0].split()
		for s in range(0,len(x)):
			print(x[s])

		print("\n\nDestination:")
		x = FOF[element][1].split()
		for s in range(0,len(x)):
			print(x[s])

		print("\n\nPort:")
		x = FOF[element][2].split()
		for s in range(0,len(x)):
			print(x[s])

		print("\n\n==========================================================")

#===================TESTING===================================
# FOF = []
# first = ["GRP_CCY_ALL_USRWKS GRP_ANIKI_PRINTERS", "HOST_10.90.0.1", "TCP_443"]
# second = ["GRP_CCY_ALL_USRWKS", "HOST_10.90.2.1", "TCP_22"]
# third = ["GRP_CCY_ALL_USRWKS GRP_ANIKI_PRINTERS", "HOST_10.90.0.1", "TCP_80"]
# fourth = ["AHEEM", "HOST_10.90.0.1", "TCP_443 TCP_80"]


# #for i in range(0,n):
# #	print(i)

# FOF.append(first)
# FOF.append(second)
# FOF.append(third)
#FOF.append(fourth)
#============================================================

# SCRIP.PY
#============================================================
FOF = []
getFOF(FOF)
concatenate(FOF, len(FOF))
print(FOF)
showFOFDone(FOF)
#============================================================